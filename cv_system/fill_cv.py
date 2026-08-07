#!/usr/bin/env python3
"""
fill_cv.py  —  Füllt CV-Daten (JSON) in die Improfy-Excel-Vorlage ("Muster").

Ergebnis:
  1) Neues Tab in der Master-Arbeitsmappe  (ID-Kxxxx_Name)
  2) Einzelne, weiterleitbare Excel nur mit diesem einen Blatt
  3) Fehlende Pflichtangaben werden GELB markiert + Checkliste (im Blatt Spalte H
     und als separate .txt-Datei)

Aufruf:
  python3 fill_cv.py  kandidat.json  "ID-K0148_Max_Mustermann"

Optional:
  --master  Pfad zur Master-Arbeitsmappe  (Default: siehe MASTER unten)
  --out     Ausgabeordner für die Einzel-Excel + Checkliste (Default: ./ausgabe)
"""
import argparse
import copy
import datetime
import json
import os
import re
import shutil

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------- Konfiguration
MASTER_DEFAULT = "/Users/ikromsdikiy/Desktop/Improfy/Sahar Mohammadi Niyay Rodsary.xlsx"
TEMPLATE_SHEET = "Muster"
GELB = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FEHLT = "⚠ BITTE ERGÄNZEN"
DATA_FONT_COLOR = "FF000000"  # eingetragene Daten schwarz
DATA_FONT_SIZE = 12           # etwas größer als Vorlage-Standard
DEFAULT_KUNDE_VON = "Ikrom"   # Feld 'Kunde von:' (B3) – Standardwert
KUNDE_VON_ZELLE = "B3"

# --------------------------------------------------- FELD-MAP der Muster-Vorlage
# (feste Zeilen, weil "Muster" immer gleich aufgebaut ist)
KOPF = {  # feld -> Zelle
    "vorname": "B10",
    "nachname": "B11",
    "angestrebter_job": "B12",
    "geburtsdatum": "B13",
    "mobil": "B14",
    "email": "B15",
    "adresse": "B17",
    "fuehrerschein": "B19",
}
# Berufserfahrung: 7 Slots. Je Slot: (Zeitraum-Zelle, Firma, Jobtitel, [Tätigkeiten...])
# Der fixe Improfy-Block (Zeilen 24-28) bleibt unangetastet.
BERUF_SLOTS = [
    ("B30", "D30", "D31", ["D32", "D33", "D34"]),
    ("B38", "D38", "D39", ["D40", "D41", "D42"]),
    ("B46", "D46", "D47", ["D48", "D49", "D50"]),
    ("B55", "D55", "D56", ["D57", "D58", "D59"]),
    ("B63", "D63", "D64", ["D65", "D66", "D67"]),
    ("B70", "D70", "D71", ["D72", "D73", "D74"]),
    ("B76", "D76", "D77", ["D78", "D79"]),
]
# Bildung: 4 Slots. Je Slot: (Zeitraum-Zelle, Abschluss-Art, Institution, Note)
BILDUNG_SLOTS = [
    ("B98", "D98", "D99", "D100"),
    ("B102", "D102", "D103", "D104"),
    ("B112", "D112", "D113", "D114"),
    ("B118", "D118", "D119", "D120"),
]
ZUSATZQUAL_ZELLEN = ["D124", "D125", "D126"]
# EDV: (Programm, Sterne)   Soft Skills: (Eigenschaft, Sterne)
EDV_SLOTS = [("B84", "C84"), ("B85", "C85"), ("B86", "C86"), ("B87", "C87"), ("B88", "C88")]
SOFT_SKILL_LABEL_COL = "D"  # Labels in D84..., Sterne in E84...
SOFT_SKILL_START = 84
SOFT_SKILL_COUNT = 10        # max. Anzahl Soft Skills (Platz bis D96)
SOFT_SKILL_CLEAR_END = 96    # Vorlage hat Vorgaben bis D96/E96 -> alle leeren
# Sprachen: (Sprache, Niveau)
SPRACHE_SLOTS = [("B137", "D137"), ("B138", "D138"), ("B139", "D139"),
                 ("B140", "D140"), ("B141", "D141")]
UEBER_MICH_ZELLE = "B147"   # oberer linker Anker des Textfelds (B147:D163)
FOTO_ZELLE = "B7"

# ---------------------------------------------------------------------- Helpers
def _parse_datum(text):
    m = re.match(r"^\s*(\d{1,2})\.(\d{1,2})\.(\d{4})\s*$", str(text or ""))
    if m:
        d, mo, y = map(int, m.groups())
        try:
            return datetime.datetime(y, mo, d)
        except ValueError:
            pass
    return text or None


class Filler:
    def __init__(self, ws):
        self.ws = ws
        self.fehlend = []  # Liste von (Feldname, Zelle)

    def set(self, zelle, wert):
        self.ws[zelle] = wert

    def pflicht(self, zelle, wert, feldname):
        """Setzt Wert; markiert gelb + notiert Checkliste, wenn leer."""
        if wert not in (None, "", []):
            self.ws[zelle] = wert
        else:
            self.ws[zelle] = FEHLT
            self.ws[zelle].fill = GELB
            self.fehlend.append((feldname, zelle))

    def markiere_fehlend(self, feldname, zelle=""):
        self.fehlend.append((feldname, zelle))


def fuelle_blatt(ws, data):
    f = Filler(ws)

    # ---- Kunde von: (oben, immer ausfüllen)
    ws[KUNDE_VON_ZELLE] = data.get("kunde_von") or DEFAULT_KUNDE_VON

    # ---- Improfy-Block: Jobtitel geschlechtsrichtig (D25 fix in der Vorlage = Teilnehmerin)
    g = (data.get("geschlecht") or "").strip().lower()
    if g == "m":
        ws["D25"] = "Teilnehmer"
    elif g == "w":
        ws["D25"] = "Teilnehmerin"

    # ---- Kopf / Personendaten
    f.pflicht(KOPF["vorname"], data.get("vorname"), "Vorname")
    f.pflicht(KOPF["nachname"], data.get("nachname"), "Nachname")
    f.pflicht(KOPF["angestrebter_job"], data.get("angestrebter_job"), "Angestrebter Job")
    f.pflicht(KOPF["geburtsdatum"], _parse_datum(data.get("geburtsdatum")), "Geburtsdatum")
    f.pflicht(KOPF["mobil"], data.get("mobil"), "Mobilnummer")
    f.pflicht(KOPF["email"], data.get("email"), "E-Mail")
    f.pflicht(KOPF["adresse"], data.get("adresse"), "Adresse")

    # Nicht-EU-Felder immer leeren (sonst bleiben Vorlagen-Reste 'Ja'/'anerkannt')
    f.set("B20", None)
    f.set("D20", None)
    f.set("B21", None)
    fs = data.get("fuehrerschein") or {}
    if fs.get("vorhanden"):
        f.set(KOPF["fuehrerschein"], f"Ja | Klasse {fs.get('klasse') or '?'}")
        if not fs.get("eu", True):  # Nicht-EU: Zusatzfeld B20/D20 befüllen
            f.set("B20", f"Ja | Klasse {fs.get('klasse') or '?'} | {fs.get('land') or 'Ausland'}")
            f.set("D20", "Anerkennung prüfen")
    else:
        f.set(KOPF["fuehrerschein"], "Nein")

    # Foto-Link
    if not (data.get("foto_link")):
        f.markiere_fehlend("Foto-Link (Kundenordner/Drive)", FOTO_ZELLE)

    # ---- Berufserfahrung
    beruf = data.get("berufserfahrung") or []
    if not beruf:
        f.markiere_fehlend("Berufserfahrung (mind. 1 Station)")
    for slot, eintrag in zip(BERUF_SLOTS, beruf[: len(BERUF_SLOTS)]):
        zeit, firma, jt, taet_zellen = slot
        f.set(zeit, eintrag.get("zeitraum") or "")
        f.set(firma, eintrag.get("firma") or "")
        f.set(jt, eintrag.get("jobtitel") or "")
        for zelle, taet in zip(taet_zellen, (eintrag.get("taetigkeiten") or [])):
            f.set(zelle, taet)

    # ---- Bildung
    bildung = data.get("bildung") or []
    if not bildung:
        f.markiere_fehlend("Bildung (mind. 1 Abschluss)")
    for slot, eintrag in zip(BILDUNG_SLOTS, bildung[: len(BILDUNG_SLOTS)]):
        zeit, art, inst, note = slot
        f.set(zeit, eintrag.get("zeitraum") or "")
        f.set(art, eintrag.get("abschluss") or "")
        f.set(inst, eintrag.get("institution") or "")
        if eintrag.get("note"):
            f.set(note, eintrag["note"])

    # ---- Zusatzqualifikationen
    for zelle, zq in zip(ZUSATZQUAL_ZELLEN, (data.get("zusatzqualifikationen") or [])):
        f.set(zelle, zq)

    # ---- EDV
    for (prog_z, stern_z), edv in zip(EDV_SLOTS, (data.get("edv_kenntnisse") or [])):
        f.set(prog_z, edv.get("programm") or "")
        if edv.get("sterne"):
            f.set(stern_z, edv["sterne"])

    # ---- Soft Skills  (überschreibt die Vorgabe-Liste, falls Daten vorhanden)
    softskills = data.get("soft_skills") or []
    if softskills:
        # ERST die komplette Vorgabe-Liste der Vorlage leeren (bis D96/E96),
        # sonst bleiben Reste stehen und Skills erscheinen doppelt.
        for zeile in range(SOFT_SKILL_START, SOFT_SKILL_CLEAR_END + 1):
            f.set(f"{SOFT_SKILL_LABEL_COL}{zeile}", None)
            f.set(f"E{zeile}", None)
        for i, sk in enumerate(softskills[:SOFT_SKILL_COUNT]):
            zeile = SOFT_SKILL_START + i
            f.set(f"{SOFT_SKILL_LABEL_COL}{zeile}", sk.get("eigenschaft") or "")
            f.set(f"E{zeile}", sk.get("sterne") or 5)

    # ---- Sprachen
    sprachen = data.get("sprachen") or []
    if not any((s.get("sprache") or "").upper().startswith("DEUTSCH") for s in sprachen):
        f.markiere_fehlend("Sprache Deutsch (Niveau)")
    for (spr_z, niv_z), spr in zip(SPRACHE_SLOTS, sprachen):
        f.set(spr_z, (spr.get("sprache") or "").upper())
        f.set(niv_z, spr.get("niveau") or "")

    # ---- Über mich
    f.pflicht(UEBER_MICH_ZELLE, data.get("ueber_mich"), "Über-mich-Text")

    # ---- Hobbys
    hob_zelle = _finde_hobby_zelle(ws)
    if data.get("hobbys"):
        f.set(hob_zelle, data["hobbys"])
    else:
        f.markiere_fehlend("Hobbys", hob_zelle)

    # ---- Eingetragene Daten schwarz + etwas größer
    restyle_daten(ws, hob_zelle)

    # ---- Zellverbindungen aufheben, damit in Excel frei kopierbar
    entmerge_fuer_copy(ws)

    return f.fehlend


def entmerge_fuer_copy(ws):
    """Hebt alle verbundenen Zellen auf (sonst blockt Excel das Kopieren)."""
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    # 'Über mich'-Text lesbar halten: Umbruch an + Zeile hoch genug
    c = ws[UEBER_MICH_ZELLE]
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[int(UEBER_MICH_ZELLE[1:])].height = 320


def _daten_zellen(hob_zelle):
    """Alle Zellen, die echte Kandidaten-Daten enthalten (zum Umformatieren)."""
    cells = list(KOPF.values()) + [KUNDE_VON_ZELLE, "B20", "D20", "D24", "D25", "D26", "D27", "D28"]
    for zeit, firma, jt, taets in BERUF_SLOTS:
        cells += [zeit, firma, jt, *taets]
    for zeit, art, inst, note in BILDUNG_SLOTS:
        cells += [zeit, art, inst, note]
    cells += ZUSATZQUAL_ZELLEN
    for prog, stern in EDV_SLOTS:
        cells += [prog, stern]
    for r in range(SOFT_SKILL_START, SOFT_SKILL_CLEAR_END + 1):
        cells += [f"D{r}", f"E{r}"]
    for spr, niv in SPRACHE_SLOTS:
        cells += [spr, niv]
    cells += [UEBER_MICH_ZELLE, hob_zelle]
    return cells


def restyle_daten(ws, hob_zelle):
    """Eingetragene Daten schwarz + etwas größer (Schriftart/Fett bleiben)."""
    for zelle in _daten_zellen(hob_zelle):
        c = ws[zelle]
        if c.value in (None, ""):
            continue
        alt = c.font
        c.font = Font(name=alt.name, size=DATA_FONT_SIZE, bold=alt.bold,
                      italic=alt.italic, color=DATA_FONT_COLOR)


def schreibe_kopiertext(data, pfad):
    """Erzeugt eine sauber kopierbare Text-Datei mit allen Daten (Copy & Paste)."""
    L = []
    fs = data.get("fuehrerschein") or {}
    fs_txt = (f"Ja, Klasse {fs.get('klasse') or '?'}" if fs.get("vorhanden") else "Nein")
    L += [f"KUNDE VON: {data.get('kunde_von') or DEFAULT_KUNDE_VON}", "",
          "PERSÖNLICHE DATEN",
          f"Vorname: {data.get('vorname','')}",
          f"Nachname: {data.get('nachname','')}",
          f"Angestrebter Job: {data.get('angestrebter_job','')}",
          f"Geburtsdatum: {data.get('geburtsdatum','')}",
          f"Mobil: {data.get('mobil','')}",
          f"E-Mail: {data.get('email','')}",
          f"Adresse: {data.get('adresse','')}",
          f"Führerschein: {fs_txt}", ""]
    L.append("BERUFSERFAHRUNG")
    for b in data.get("berufserfahrung", []):
        L.append(f"{b.get('zeitraum','')} — {b.get('firma','')}")
        if b.get("jobtitel"):
            L.append(f"  {b['jobtitel']}")
        for t in b.get("taetigkeiten", []):
            L.append(f"  - {t}")
        L.append("")
    L.append("BILDUNG")
    for e in data.get("bildung", []):
        L.append(f"{e.get('zeitraum','')} — {e.get('abschluss','')}"
                 + (f" ({e['institution']})" if e.get("institution") else ""))
    L.append("")
    if data.get("zusatzqualifikationen"):
        L.append("ZUSATZQUALIFIKATIONEN")
        L += [f"- {z}" for z in data["zusatzqualifikationen"]]
        L.append("")
    L.append("SPRACHEN")
    L += [f"- {s.get('sprache','')}: {s.get('niveau','')}" for s in data.get("sprachen", [])]
    L.append("")
    if data.get("edv_kenntnisse"):
        L.append("EDV / IT")
        L += [f"- {k.get('programm','')} ({k.get('sterne','')}/5)" for k in data["edv_kenntnisse"]]
        L.append("")
    if data.get("soft_skills"):
        L.append("SOFT SKILLS")
        L += [f"- {k.get('eigenschaft','')} ({k.get('sterne','')}/5)" for k in data["soft_skills"]]
        L.append("")
    L += ["ÜBER MICH", data.get("ueber_mich", ""), ""]
    if data.get("hobbys"):
        L += ["HOBBYS", data["hobbys"], ""]
    with open(pfad, "w", encoding="utf-8") as fh:
        fh.write("\n".join(L))


def _finde_hobby_zelle(ws):
    """Sucht das 'Hob...'-Label in Spalte A und gibt die Wert-Zelle daneben (B) zurück."""
    for row in ws.iter_rows(min_col=1, max_col=1, max_row=ws.max_row):
        c = row[0]
        if isinstance(c.value, str) and c.value.strip().lower().startswith("hob"):
            return f"B{c.row}"
    return "B214"  # Fallback (wie in bestehenden Blättern)


def schreibe_checkliste_ins_blatt(ws, fehlend):
    ws["H1"] = "⚠ FEHLENDE / ZU ERGÄNZENDE ANGABEN"
    ws["H1"].fill = GELB
    if not fehlend:
        ws["H2"] = "✓ Alle Pflichtangaben vorhanden."
        return
    for i, (feld, zelle) in enumerate(fehlend, start=2):
        ort = f"  (Zelle {zelle})" if zelle else ""
        ws[f"H{i}"] = f"• {feld}{ort}"


# ------------------------------------------------------------------------ Main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("json_datei", help="Kandidaten-JSON (siehe schema.py)")
    ap.add_argument("blatt_name", help='z.B. "ID-K0148_Max_Mustermann"')
    ap.add_argument("--master", default=MASTER_DEFAULT)
    ap.add_argument("--out", default=os.path.join(os.path.dirname(__file__), "ausgabe"))
    args = ap.parse_args()

    with open(args.json_datei, encoding="utf-8") as fh:
        data = json.load(fh)
    os.makedirs(args.out, exist_ok=True)

    # 1) Master laden + Muster klonen
    wb = openpyxl.load_workbook(args.master)
    if TEMPLATE_SHEET not in wb.sheetnames:
        raise SystemExit(f"Vorlagenblatt '{TEMPLATE_SHEET}' nicht gefunden.")
    if args.blatt_name in wb.sheetnames:  # Re-Run: altes Blatt entfernen
        del wb[args.blatt_name]
    neu = wb.copy_worksheet(wb[TEMPLATE_SHEET])
    neu.title = args.blatt_name

    fehlend = fuelle_blatt(neu, data)
    schreibe_checkliste_ins_blatt(neu, fehlend)

    # Neues Blatt nach vorne (Position 2, direkt hinter der ersten Übersicht)
    wb.move_sheet(args.blatt_name, -(len(wb.sheetnames) - 2))
    wb.save(args.master)

    # 2) Einzelne, weiterleitbare Excel  (nur dieses eine Blatt)
    einzel_pfad = os.path.join(args.out, f"{args.blatt_name}.xlsx")
    shutil.copy(args.master, einzel_pfad)
    wb2 = openpyxl.load_workbook(einzel_pfad)
    for name in list(wb2.sheetnames):
        if name != args.blatt_name:
            del wb2[name]
    wb2.active = 0
    wb2.save(einzel_pfad)

    # 3) Checkliste als Text
    txt_pfad = os.path.join(args.out, f"{args.blatt_name}_CHECKLISTE.txt")
    with open(txt_pfad, "w", encoding="utf-8") as fh:
        fh.write(f"CHECKLISTE – {args.blatt_name}\n{'='*50}\n")
        if fehlend:
            fh.write("Folgende Angaben fehlen im CV und müssen ergänzt werden:\n\n")
            for feld, zelle in fehlend:
                fh.write(f"  • {feld}" + (f"  (Zelle {zelle})\n" if zelle else "\n"))
        else:
            fh.write("✓ Alle Pflichtangaben vorhanden.\n")

    # 4) Kopierbare Text-Datei (Copy & Paste)
    txt2 = os.path.join(args.out, f"{args.blatt_name}_KOPIERBAR.txt")
    schreibe_kopiertext(data, txt2)

    # Report
    print(f"✓ Kopierbarer Text:      {txt2}")
    print(f"✓ Master aktualisiert:   {args.master}")
    print(f"✓ Einzel-Excel:          {einzel_pfad}")
    print(f"✓ Checkliste:            {txt_pfad}")
    if fehlend:
        print(f"\n⚠ {len(fehlend)} fehlende Angabe(n):")
        for feld, zelle in fehlend:
            print(f"   • {feld}" + (f"  ({zelle})" if zelle else ""))
    else:
        print("\n✓ Keine fehlenden Pflichtangaben.")


if __name__ == "__main__":
    main()
