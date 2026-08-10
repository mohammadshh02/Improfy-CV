#!/usr/bin/env python3
"""
Selbsttest
==========
Prüft nach einer Änderung, ob die App noch das tut, was sie soll — Anmeldung,
Excel, Foto-Aufbereitung, PDF-Design und Admin-Panel.

    python3 selbsttest.py            # alles ohne KI (kostet nichts)
    python3 selbsttest.py --mit-ki   # zusätzlich der Schnell-Modus (API-Guthaben!)

Braucht kein Passwort: der Test setzt die Sitzung direkt. Dass die Anmeldung
selbst funktioniert, wird über den umgekehrten Weg geprüft — falsches Passwort
und Zugriff ohne Anmeldung müssen scharf abgewiesen werden.

Rückgabewert 0 = alles bestanden, 1 = mindestens ein Punkt fehlgeschlagen.
"""
import base64
import io
import json
import os
import random
import sys

HIER = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HIER)
os.chdir(HIER)

import server  # noqa: E402

server.app.config["TESTING"] = True

_ergebnisse = []


def pruefe(text, bedingung, zusatz=""):
    _ergebnisse.append(bool(bedingung))
    marke = "ok    " if bedingung else "FEHLER"
    print(f"  [{marke}] {text}" + (f" — {zusatz}" if zusatz else ""))
    return bool(bedingung)


def angemeldet(kennung="koeln", rolle="stadt"):
    """Client mit gesetzter Sitzung, ohne das Passwort zu kennen."""
    c = server.app.test_client()
    with c.session_transaction() as s:
        s["kennung"], s["rolle"] = kennung, rolle
    return c


TESTDATEN = {
    "vorname": "Max", "nachname": "Mustermann", "geschlecht": "männlich",
    "angestrebter_job": "Lagerhelfer", "geburtsdatum": "01.01.1990",
    "mobil": "0170 0000000", "email": "max@example.org",
    "adresse": "Musterweg 1, 50667 Köln",
    "berufserfahrung": [{"zeitraum": "2020-2024", "firma": "Beispiel GmbH",
                         "jobtitel": "Lagerist", "taetigkeiten": ["Kommissionierung"]}],
    "bildung": [{"zeitraum": "2006-2010", "abschluss": "Hauptschulabschluss",
                 "institution": "Musterschule", "note": ""}],
    "sprachen": [{"sprache": "Deutsch", "niveau": 5}],
    "edv_kenntnisse": [{"programm": "MS Office", "sterne": 4}],
    "soft_skills": [{"eigenschaft": "Teamfähigkeit", "sterne": 5}],
    "zusatzqualifikationen": ["Staplerschein"], "hobbys": "Fußball",
    "ueber_mich": "Testdatensatz zur Funktionsprüfung.",
    "fuehrerschein": {"vorhanden": True, "klasse": "B", "eu": True},
}


def test_absperrung():
    print("\nAbsperrung")
    c = server.app.test_client()
    r = c.get("/")
    pruefe("Ohne Anmeldung führt / zur Loginseite", r.status_code == 302
           and "/login" in (r.headers.get("Location") or ""))
    r = c.post("/generate", data={"vorname": "X"})
    pruefe("Laufende Anfragen bekommen 401 als JSON, nicht die Login-Seite",
           r.status_code == 401 and r.is_json)
    r = c.post("/login", data={"kennung": "koeln", "passwort": "bestimmt-falsch"})
    pruefe("Falsches Passwort wird abgewiesen",
           "Passwort stimmt nicht" in r.get_data(as_text=True))
    r = angemeldet().get("/admin")
    pruefe("Ein Standort kommt nicht ins Admin-Panel", r.status_code == 302)


def test_excel():
    print("\nExcel")
    import openpyxl
    formular = {"vorname": "Max", "nachname": "Mustermann", "geschlecht": "männlich",
                "angestrebter_job": "Lagerhelfer", "geburtsdatum": "01.01.1990"}
    r = angemeldet().post("/generate", data=formular)
    if not pruefe("Excel wird erzeugt", r.status_code == 200, f"{len(r.data)} Bytes"):
        return
    antwort = r.get_json()
    roh = base64.b64decode(antwort["file_b64"])
    ws = openpyxl.load_workbook(io.BytesIO(roh))[antwort["filename"].rsplit(".", 1)[0][:31]]
    pruefe("Keine verbundenen Zellen (sonst nicht kopierbar)",
           len(ws.merged_cells.ranges) == 0, f"{len(ws.merged_cells.ranges)} gefunden")
    pruefe("Kein Blattschutz", not ws.protection.sheet)
    farbe = ws["B10"].font.color.rgb if ws["B10"].font.color else None
    pruefe("Daten stehen in schwarzer Schrift", farbe == "FF000000", str(farbe))


def test_foto():
    print("\nFoto-Aufbereitung")
    from PIL import Image

    # Handyfoto: quer aufgenommen, per EXIF als gedreht markiert.
    quer = Image.new("RGB", (400, 200), (18, 161, 80))
    exif = quer.getexif()
    exif[0x0112] = 6
    puffer = io.BytesIO()
    quer.save(puffer, "JPEG", exif=exif, quality=90)
    ergebnis = Image.open(io.BytesIO(base64.b64decode(
        server._foto_data_uri(puffer.getvalue(), "image/jpeg").split(",", 1)[1])))
    pruefe("EXIF-Drehung wird fest ins Bild gerechnet",
           ergebnis.size == (200, 400), f"{quer.size} -> {ergebnis.size}")

    # Fotoähnliches PNG mit Transparenz: muss auf Weiß landen, nicht auf Schwarz.
    random.seed(1)
    p = Image.new("RGBA", (700, 700))
    px = p.load()
    for x in range(700):
        for y in range(700):
            px[x, y] = ((0, 0, 0, 0) if x < 100 and y < 100 else
                        (random.randrange(256), random.randrange(256), random.randrange(256), 255))
    puffer = io.BytesIO()
    p.save(puffer, format="PNG")
    aus = Image.open(io.BytesIO(base64.b64decode(
        server._foto_data_uri(puffer.getvalue(), "image/png").split(",", 1)[1])))
    if aus.format == "JPEG":
        ecke = aus.convert("RGB").getpixel((20, 20))
        pruefe("Transparenz wird weiß hinterlegt, nicht schwarz",
               all(k > 235 for k in ecke), str(ecke))
    else:
        pruefe("Kleines Bild bleibt unverändert (bewusst)", True, aus.format)


def test_pdf():
    print("\nPDF-Design")
    if not server.CHROME:
        pruefe("Chrome vorhanden", False, "kein Chrome gefunden — PDF nicht möglich")
        return
    r = angemeldet().post("/cv-pdf", data={"daten": json.dumps(TESTDATEN), "design": "gruen"})
    pruefe("PDF wird erzeugt", r.status_code == 200 and r.data[:4] == b"%PDF",
           f"{len(r.data)} Bytes")


def test_panel():
    print("\nAdmin-Panel")
    seite = angemeldet("admin", "admin").get("/admin").get_data(as_text=True)
    for begriff in ("Nutzungs-Übersicht", "Nutzung pro Tag", "Standorte im Vergleich",
                    "System & Sicherheit", "Letzte Ereignisse"):
        pruefe(f"Abschnitt „{begriff}“ ist da", begriff.replace("&", "&amp;") in seite)


def test_ki():
    print("\nKI-Auswertung (verbraucht API-Guthaben)")
    text = ("Erika Beispiel, geboren am 12.03.1988, Teststrasse 5, 68159 Mannheim. "
            "Sucht Stelle als Bueroassistenz. 2015 bis 2023 Sachbearbeiterin bei der "
            "Muster AG: Rechnungspruefung, Terminplanung. Realschulabschluss 2004. "
            "Deutsch Muttersprache, Englisch gut. Word, Excel, Outlook.")
    if not server.ki_verfuegbar():
        pruefe("KI erreichbar", False, "weder Claude-CLI noch API-Key")
        return
    r = angemeldet().post("/generate-auto", data={"rohtext": text})
    if not pruefe("Schnell-Modus liefert eine Excel", r.status_code == 200):
        print("   ", r.get_data(as_text=True)[:200])
        return
    d = r.get_json()
    daten = d.get("daten") or {}
    pruefe("Name richtig erkannt", daten.get("vorname") == "Erika",
           f"{daten.get('vorname')} {daten.get('nachname')}")
    pruefe("Berufserfahrung übernommen", len(daten.get("berufserfahrung") or []) >= 1,
           f"{len(daten.get('berufserfahrung') or [])} Einträge")


if __name__ == "__main__":
    print("Selbsttest der Improfy CV-App")
    test_absperrung()
    test_excel()
    test_foto()
    test_pdf()
    test_panel()
    if "--mit-ki" in sys.argv:
        test_ki()
    else:
        print("\n(KI-Test übersprungen — mit --mit-ki einschalten)")

    fehler = _ergebnisse.count(False)
    print(f"\n{len(_ergebnisse) - fehler} von {len(_ergebnisse)} bestanden.")
    sys.exit(1 if fehler else 0)
